import boto3
import openpyxl
import time
import logging

# Configure logging
logging.basicConfig(filename='dns_management.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_or_get_hosted_zone(client, hosted_zone_name):
    try:
        # Try to get the hosted zone ID
        response = client.list_hosted_zones_by_name(DNSName=hosted_zone_name)
        hosted_zone_id = response['HostedZones'][0]['Id']
    except client.exceptions.ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchHostedZone':
            # Hosted zone doesn't exist, create a new one
            logging.info(f'Creating new hosted zone for {hosted_zone_name}')
            response = client.create_hosted_zone(
                Name=hosted_zone_name,
                CallerReference=str(time.time())  # Ensure a unique reference
                # Add any other required parameters
            )
            hosted_zone_id = response['HostedZone']['Id']
        else:
            # An unexpected error occurred
            logging.error(f'Error while creating/getting hosted zone: {e}')
            raise

    return hosted_zone_id

client = boto3.client('route53')

workbook = openpyxl.load_workbook('dnsentries.xlsx')
print("workbook is opened")
worksheet = workbook.active

dnslist = []
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
    dnsdict = dict(zip(worksheet[1], row))
    dnslist.append(dnsdict)

for routelist in dnslist:
    isupdated = False
    # Check if the hosted zone exists or create a new one
    hosted_zone_id = create_or_get_hosted_zone(client, routelist['hostedzonename'])
    
    # Continue with the existing record management logic...
    sets = client.list_resource_record_sets(HostedZoneId=hosted_zone_id)
    hostname = routelist['recordname']
    for rset in sets['ResourceRecordSets']:
        curttl = rset['TTL']
        if hostname[-1:] != ".":
            hostname += "."
        if rset['Name'] == hostname and rset['Type'] == routelist['recordtype']:
            # UPSERT the record
            logging.info(f'Updating {hostname}')
            client.change_resource_record_sets(
                HostedZoneId=hosted_zone_id,
                ChangeBatch={
                    'Changes': [
                        {
                            'Action': 'UPSERT',
                            'ResourceRecordSet': {
                                'Name': hostname,
                                'Type': routelist['recordtype'],
                                'TTL': curttl,
                                'ResourceRecords': [
                                    {
                                        'Value': routelist['value']
                                    }
                                ]
                            }
                        }
                    ]
                }
            )
            isupdated = True
        else:
            if rset['Name'] == hostname and rset['Type'] != routelist['recordtype']:
                # DELETE the record
                logging.info(f'Deleting {hostname}')
                logging.info(f'Existing value: {rset["ResourceRecords"][0]["Value"]}')
                client.change_resource_record_sets(
                    HostedZoneId=hosted_zone_id,
                    ChangeBatch={
                        'Changes': [
                            {
                                'Action': 'DELETE',
                                'ResourceRecordSet': {
                                    'Name': hostname,
                                    'Type': rset['Type'],
                                    'TTL': curttl,
                                    'ResourceRecords': [
                                        {
                                            'Value': rset['ResourceRecords'][0]['Value']
                                        }
                                    ]
                                }
                            }
                        ]
                    }
                )
                # INSERT the record
                logging.info(f'Inserting {hostname}')
                client.change_resource_record_sets(
                    HostedZoneId=hosted_zone_id,
                    ChangeBatch={
                        'Changes': [
                            {
                                'Action': 'CREATE',
                                'ResourceRecordSet': {
                                    'Name': hostname,
                                    'Type': routelist['recordtype'],
                                    'TTL': 300,
                                    'ResourceRecords': [
                                        {
                                            'Value': routelist['value']
                                        }
                                    ]
                                }
                            }
                        ]
                    }
                )
                isupdated = True
    if not isupdated:
        # INSERT the record
        logging.info(f'Inserting {hostname}')
        client.change_resource_record_sets(
            HostedZoneId=hosted_zone_id,
            ChangeBatch={
                'Changes': [
                    {
                        'Action': 'CREATE',
                        'ResourceRecordSet': {
                            'Name': hostname,
                            'Type': routelist['recordtype'],
                            'TTL': 300,
                            'ResourceRecords': [
                                {
                                    'Value': routelist['value']
                                }
                            ]
                        }
                    }
                ]
            }
        )
